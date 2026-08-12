import unittest
from pathlib import Path

from openpyxl import load_workbook

from report_generator import ReportGenerator


class TestExcelReportFormatting(unittest.TestCase):
    def test_generate_xlsx_report_formats_title_and_panes(self):
        analysis = {
            "old_folder_name": "old",
            "new_folder_name": "new",
            "timestamp": "2026-08-12 12:00:00",
            "results": [
                {
                    "Comparison_Status": "MODIFIED",
                    "Changed_Component": "Order_Date",
                    "Difference_Details": ["Order_Date changed from '2024-01-01' to '2024-01-02'"],
                    "Old_Order_ID": "ORD-1",
                    "New_Order_ID": "ORD-1",
                    "Old_Order_Date": "2024-01-01",
                    "New_Order_Date": "2024-01-02",
                    "Old_Customer_Name": "Alpha",
                    "New_Customer_Name": "Alpha",
                    "Old_City": "A",
                    "New_City": "A",
                    "Old_State": "S",
                    "New_State": "S",
                    "Old_Region": "R",
                    "New_Region": "R",
                    "Old_Country": "US",
                    "New_Country": "US",
                    "Old_Category": "Cat",
                    "New_Category": "Cat",
                    "Old_Sub_Category": "Sub",
                    "New_Sub_Category": "Sub",
                    "Old_Product_Name": "Widget",
                    "New_Product_Name": "Widget",
                }
            ],
        }

        path = Path("tmp_report.xlsx")
        ReportGenerator().generate_xlsx_report(analysis, path)

        try:
            wb = load_workbook(path)
            ws = wb.active

            self.assertEqual(ws.freeze_panes, "B5")
            self.assertGreater(len(ws.merged_cells.ranges), 0)
            self.assertTrue(str(ws["A1"].value).startswith("Folder Difference Analysis:"))
            self.assertEqual(ws["A4"].fill.start_color.rgb, "001F497D")
            self.assertEqual(ws["A5"].alignment.vertical, "top")
            self.assertTrue(ws["A5"].alignment.wrap_text)
        finally:
            if path.exists():
                path.unlink()


if __name__ == "__main__":
    unittest.main()
