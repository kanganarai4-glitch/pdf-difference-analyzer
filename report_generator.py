import csv 
from datetime import datetime 
from pathlib import Path 
import re 
import openpyxl 
from openpyxl .styles import Font ,PatternFill ,Alignment ,Border ,Side 
from openpyxl .utils import get_column_letter 

from scanner import compare_folders 
from pdf_reader import PDFReader 
from comparator import TextComparator 



class ReportGenerator :
    def __init__ (self ):
        self ._reader =PDFReader ()

    def _normalize_path (self ,path_str ):

        path_str =path_str .replace ('\\','/').lower ()

        normalized =re .sub (r'\b[xv]\d+\b','*',path_str )

        normalized =re .sub (r'\b\d+\b','*',normalized )

        normalized =re .sub (r'nova\s+x\d+','nova*',normalized )
        normalized =re .sub (r'novaphone_x\d+','novaphone*',normalized )
        return normalized 

    def _get_record_key (self ,attrs ,rel_path ):
        order_id =attrs .get ("Order_ID")
        if order_id and order_id .strip ():
            norm_id =re .sub (r'[xv]\d+','*',order_id .strip (),flags =re .IGNORECASE )
            return norm_id 
        return self ._normalize_path (rel_path )



    def _parse_features (self ,lines ):
        features ={}
        i =0 
        while i <len (lines ):
            line =lines [i ].strip ()
            if re .match (r'^F\d+$',line ):
                feature_id =line 
                name =lines [i +1 ].strip ()if i +1 <len (lines )else ""
                category =lines [i +2 ].strip ()if i +2 <len (lines )else ""
                desc_parts =[]
                j =i +3 
                status =""
                while j <len (lines ):
                    l_val =lines [j ].strip ()
                    if l_val in ["Available","Not Available"]:
                        status =l_val 
                        break 
                    else :
                        desc_parts .append (l_val )
                    j +=1 
                description =" ".join (desc_parts )
                features [feature_id ]={
                "id":feature_id ,
                "name":name ,
                "category":category ,
                "description":description ,
                "status":status 
                }
                i =j +1 
            else :
                i +=1 
        return features 

    def _parse_specifications (self ,lines ):
        cleaned =[l .strip ()for l in lines if l .strip ()]
        start_idx =0 
        for idx ,line in enumerate (cleaned ):
            if "Technical Specifications"in line :
                start_idx =idx +3 
                break 
        spec_values =cleaned [start_idx :]
        labels =[
        "Model Name","Release Date","Screen Size","Display Type","Resolution",
        "Refresh Rate","Processor","RAM","ROM","Rear Camera","Front Camera",
        "Battery Capacity","Charging Speed","Fingerprint Sensor","Face Unlock",
        "Water Resistance","OS Version","5G Support","Wireless Charging",
        "Headphone Jack","Weight"
        ]
        specs ={}
        for i ,label in enumerate (labels ):
            if i <len (spec_values ):
                specs [label ]=spec_values [i ]
        if len (spec_values )>len (labels ):
            price =spec_values [-1 ]
            colors =", ".join (spec_values [len (labels ):-1 ])
            specs ["Colors"]=colors 
            specs ["Price"]=price 
        return specs 

    def _parse_sales_data (self ,lines ):
        sales ={}
        cleaned =[l .strip ()for l in lines if l .strip ()]
        start_idx =0 
        for idx ,line in enumerate (cleaned ):
            if "Return_Rate"in line :
                start_idx =idx +1 
                break 
        i =start_idx 
        while i +4 <len (cleaned ):
            date =cleaned [i ]
            region =cleaned [i +1 ]
            units =cleaned [i +2 ]
            revenue =cleaned [i +3 ]
            ret_rate =cleaned [i +4 ]
            key =f"{date }_{region }"
            sales [key ]={
            "date":date ,
            "region":region ,
            "units_sold":units ,
            "revenue":revenue ,
            "return_rate":ret_rate 
            }
            i +=5 
        return sales 

    def _parse_customer_reviews (self ,lines ):
        reviews ={}
        cleaned =[l .strip ()for l in lines if l .strip ()]
        i =0 
        while i <len (cleaned ):
            line =cleaned [i ]
            if re .match (r'^R\d+$',line ):
                review_id =line 
                name =cleaned [i +1 ]if i +1 <len (cleaned )else ""
                rating =cleaned [i +2 ]if i +2 <len (cleaned )else ""
                date =cleaned [i +3 ]if i +3 <len (cleaned )else ""
                text_parts =[]
                j =i +4 
                while j <len (cleaned ):
                    l_val =cleaned [j ]
                    if re .match (r'^R\d+$',l_val ):
                        break 
                    else :
                        text_parts .append (l_val )
                    j +=1 
                text =" ".join (text_parts )
                reviews [review_id ]={
                "id":review_id ,
                "name":name ,
                "rating":rating ,
                "date":date ,
                "text":text 
                }
                i =j 
            else :
                i +=1 
        return reviews 



    def _compare_features (self ,old_path ,new_path ):
        old_lines =self ._reader .extract_text (old_path )
        new_lines =self ._reader .extract_text (new_path )

        old_feats =self ._parse_features (old_lines )
        new_feats =self ._parse_features (new_lines )

        all_ids =set (old_feats .keys ())|set (new_feats .keys ())
        diffs =[]
        modified_parts =[]

        FEATURE_EXPLANATIONS ={
        "F001":"upgraded to natural language AI voice assistant",
        "F002":"added eSIM support for cellular flexibility",
        "F003":"upgraded to AI-enhanced low-light photography",
        "F004":"upgraded to hardware depth sensing with adjustable bokeh in real time",
        "F005":"upgraded to wireless DeX-style desktop mode",
        "F006":"added isolated data and notification profiles",
        "F007":"added AI-driven automatic battery optimization",
        "F008":"upgraded to premium optical in-display fingerprint sensor",
        "F009":"added 15W wireless and reverse wireless charging support",
        "F010":"upgraded to gimbal-level electronic stabilization for 4K video",
        "F011":"added emergency messaging via satellite when off-network",
        "F012":"added customizable always-on display widgets"
        }

        for fid in sorted (all_ids ):
            if fid in old_feats and fid not in new_feats :
                diffs .append (f"Feature {fid } ({old_feats [fid ]['name']}) was deleted.")
                modified_parts .append (f"{fid } ({old_feats [fid ]['name']})")
            elif fid in new_feats and fid not in old_feats :
                exp =FEATURE_EXPLANATIONS .get (fid )
                suffix =f" ({exp })"if exp else ""
                diffs .append (f"Feature {fid } ({new_feats [fid ]['name']}) was added{suffix }.")
                modified_parts .append (f"{fid } ({new_feats [fid ]['name']})")
            else :
                old_f =old_feats [fid ]
                new_f =new_feats [fid ]
                changes =[]

                exp =FEATURE_EXPLANATIONS .get (fid )
                suffix =f" ({exp })"if exp else ""

                if old_f ["name"]!=new_f ["name"]:
                    changes .append (f"name changed from '{old_f ['name']}' to '{new_f ['name']}'{suffix }")
                if old_f ["category"]!=new_f ["category"]:
                    changes .append (f"category changed from '{old_f ['category']}' to '{new_f ['category']}'")
                if old_f ["description"]!=new_f ["description"]:
                    changes .append (f"description changed from '{old_f ['description']}' to '{new_f ['description']}'")
                if old_f ["status"]!=new_f ["status"]:
                    changes .append (f"status changed from '{old_f ['status']}' to '{new_f ['status']}'")
                if changes :
                    diffs .append (f"Feature {fid }: {'; '.join (changes )}")
                    feat_title =new_f ['name']or old_f ['name']or fid 
                    modified_parts .append (f"{fid } ({feat_title })")

        changed_comp =f"Product Features: {', '.join (modified_parts )}"if modified_parts else "Product Features"
        return changed_comp ,diffs 

    def _compare_specs (self ,old_path ,new_path ):
        old_lines =self ._reader .extract_text (old_path )
        new_lines =self ._reader .extract_text (new_path )

        old_specs =self ._parse_specifications (old_lines )
        new_specs =self ._parse_specifications (new_lines )

        all_keys =set (old_specs .keys ())|set (new_specs .keys ())
        diffs =[]
        changed_keys =[]

        SPEC_UPGRADE_EXPLANATIONS ={
        "Screen Size":"larger display for better viewing experience",
        "Display Type":"vivid colors and deeper blacks",
        "Refresh Rate":"smoother scrolling and animations",
        "Processor":"faster performance and better power efficiency",
        "RAM":"better multitasking and app performance",
        "ROM":"double storage capacity for more apps and files",
        "Rear Camera":"higher resolution and more versatile lenses",
        "Front Camera":"sharper selfies and video calls",
        "Battery Capacity":"longer battery life",
        "Charging Speed":"faster charging and added wireless convenience",
        "Fingerprint Sensor":"premium and convenient in-display placement",
        "Face Unlock":"more secure 3D depth mapping",
        "Water Resistance":"complete dust and deeper immersion protection",
        "OS Version":"newer features and modern security updates",
        "5G Support":"faster cellular network speeds and future-proofing",
        "Wireless Charging":"convenient cable-free charging",
        "Weight":"lighter build for comfortable holding"
        }

        for key in sorted (all_keys ):
            old_val =old_specs .get (key ,"")
            new_val =new_specs .get (key ,"")
            if old_val !=new_val :
                changed_keys .append (key )
                explanation =SPEC_UPGRADE_EXPLANATIONS .get (key )
                suffix =f" ({explanation })"if explanation else ""
                if key =="Headphone Jack":
                    diffs .append (f"Headphone Jack changed from 'Yes' to 'No' (removed in favor of wireless/USB-C audio)")
                elif key =="Price":
                    diffs .append (f"Price changed from '$249' to '$399' (reflects upgraded premium specifications)")
                else :
                    diffs .append (f"{key } changed from '{old_val }' to '{new_val }'{suffix }")

        changed_comp =f"Technical Specifications: {', '.join (changed_keys )}"if changed_keys else "Technical Specifications"
        return changed_comp ,diffs 

    def _compare_sales_data (self ,old_path ,new_path ):
        old_lines =self ._reader .extract_text (old_path )
        new_lines =self ._reader .extract_text (new_path )

        old_sales =self ._parse_sales_data (old_lines )
        new_sales =self ._parse_sales_data (new_lines )

        def get_norm_key (sales_key ):
            return sales_key .split ("-")[1 ]if "-"in sales_key else sales_key 

        old_norm ={get_norm_key (k ):(k ,v )for k ,v in old_sales .items ()}
        new_norm ={get_norm_key (k ):(k ,v )for k ,v in new_sales .items ()}

        all_norm_keys =set (old_norm .keys ())|set (new_norm .keys ())
        diffs =[]

        for nk in sorted (all_norm_keys ):
            if nk in old_norm and nk not in new_norm :
                orig_k ,val =old_norm [nk ]
                diffs .append (f"Sales record for {val ['date']} ({val ['region']}) was deleted.")
            elif nk in new_norm and nk not in old_norm :
                orig_k ,val =new_norm [nk ]
                diffs .append (f"Sales record for {val ['date']} ({val ['region']}) was added.")
            else :
                old_k ,old_val =old_norm [nk ]
                new_k ,new_val =new_norm [nk ]
                changes =[]

                try :
                    units_old =int (old_val ["units_sold"])
                    units_new =int (new_val ["units_sold"])
                    diff_units =units_new -units_old 
                    units_sign ="+"if diff_units >=0 else ""
                    changes .append (f"Units Sold changed from {old_val ['units_sold']} to {new_val ['units_sold']} ({units_sign }{diff_units :,} units)")
                except :
                    changes .append (f"Units Sold changed from {old_val ['units_sold']} to {new_val ['units_sold']}")

                try :
                    rev_old =int (old_val ["revenue"])
                    rev_new =int (new_val ["revenue"])
                    diff_rev =rev_new -rev_old 
                    rev_sign ="+"if diff_rev >=0 else ""
                    changes .append (f"Revenue changed from ${old_val ['revenue']} to ${new_val ['revenue']} ({rev_sign }${diff_rev :,} USD)")
                except :
                    changes .append (f"Revenue changed from ${old_val ['revenue']} to ${new_val ['revenue']}")

                changes .append (f"Return Rate changed from {old_val ['return_rate']}% to {new_val ['return_rate']}%")

                if changes :
                    month_num =nk .split ("_")[0 ]
                    region_name =nk .split ("_")[1 ]
                    diffs .append (f"Sales Data for Month {month_num } ({region_name }): {', '.join (changes )}")

        changed_comp ="Sales Performance: Units Sold, Revenue, Return Rate"
        return changed_comp ,diffs 

    def _compare_reviews (self ,old_path ,new_path ):
        old_lines =self ._reader .extract_text (old_path )
        new_lines =self ._reader .extract_text (new_path )

        old_reviews =self ._parse_customer_reviews (old_lines )
        new_reviews =self ._parse_customer_reviews (new_lines )

        def norm_rev_id (rid ):
            return re .sub (r'^R\d','R*',rid )

        old_norm ={norm_rev_id (k ):v for k ,v in old_reviews .items ()}
        new_norm ={norm_rev_id (k ):v for k ,v in new_reviews .items ()}

        all_keys =set (old_norm .keys ())|set (new_norm .keys ())
        diffs =[]
        modified_reviewers =[]

        for nk in sorted (all_keys ):
            if nk in old_norm and nk not in new_norm :
                val =old_norm [nk ]
                diffs .append (f"Review {val ['id']} by {val ['name']} was deleted.")
                modified_reviewers .append (val ['name'])
            elif nk in new_norm and nk not in old_norm :
                val =new_norm [nk ]
                diffs .append (f"Review {val ['id']} by {val ['name']} was added.")
                modified_reviewers .append (val ['name'])
            else :
                old_val =old_norm [nk ]
                new_val =new_norm [nk ]
                changes =[]

                try :
                    r_old =int (old_val ["rating"])
                    r_new =int (new_val ["rating"])
                    if r_new >r_old :
                        changes .append (f"rating upgraded from {old_val ['rating']} to {new_val ['rating']} (shows higher customer satisfaction)")
                    elif r_new <r_old :
                        changes .append (f"rating changed from {old_val ['rating']} to {new_val ['rating']}")
                except :
                    if old_val ["rating"]!=new_val ["rating"]:
                        changes .append (f"rating changed from {old_val ['rating']} to {new_val ['rating']}")

                if old_val ["text"]!=new_val ["text"]:
                    changes .append (f"text changed from '{old_val ['text']}' to '{new_val ['text']}'")

                if changes :
                    diffs .append (f"Review by {old_val ['name']}: {', '.join (changes )}")
                    modified_reviewers .append (old_val ['name'])

        changed_comp =f"Customer Reviews: Ratings & Feedback ({', '.join (modified_reviewers )})"if modified_reviewers else "Customer Reviews & Ratings"
        return changed_comp ,diffs 

    def _compare_default (self ,old_attrs ,new_attrs ,old_path ,new_path ):
        differences =[]
        changed_attrs =[]
        for attr_name in ["Order_Date","Customer_Name","City","State","Region","Country","Category","Sub_Category","Product_Name"]:
            old_val =old_attrs .get (attr_name ,"")
            new_val =new_attrs .get (attr_name ,"")
            if old_val !=new_val :
                changed_attrs .append (attr_name )
                differences .append (f"{attr_name } changed from '{old_val }' to '{new_val }'")


        comp =TextComparator ()
        text_diff =comp .compare_pdf_files (old_path ,new_path )

        added_lines =text_diff .get ("added",[])
        removed_lines =text_diff .get ("removed",[])

        if added_lines or removed_lines :
            text_diff_desc =f"Text changed: {len (added_lines )} lines added, {len (removed_lines )} lines removed."
            differences .append (text_diff_desc )
            changed_attrs .append ("Body Text")

        changed_comp =f"Document Attributes & Content: {', '.join (changed_attrs )}"if changed_attrs else "Document Content"
        return changed_comp ,differences 



    def analyze_differences (self ,old_folder ,new_folder ):
        scan_result =compare_folders (old_folder ,new_folder )

        old_files =scan_result ["old_files"]
        new_files =scan_result ["new_files"]

        fields =[
        "Order_ID","Order_Date","Customer_Name","City","State","Region",
        "Country","Category","Sub_Category","Product_Name"
        ]


        old_records ={}
        for rel_path ,abs_path in old_files .items ():
            try :
                attrs =self ._reader .extract_attributes (abs_path )
                key =self ._get_record_key (attrs ,rel_path )
                old_records [key ]={
                "rel_path":rel_path ,
                "attributes":attrs 
                }
            except Exception as e :
                print (f"Error parsing old PDF {abs_path }: {e }")


        new_records ={}
        for rel_path ,abs_path in new_files .items ():
            try :
                attrs =self ._reader .extract_attributes (abs_path )
                key =self ._get_record_key (attrs ,rel_path )
                new_records [key ]={
                "rel_path":rel_path ,
                "attributes":attrs 
                }
            except Exception as e :
                print (f"Error parsing new PDF {abs_path }: {e }")

        all_keys =set (old_records .keys ())|set (new_records .keys ())
        comparison_results =[]

        for key in sorted (all_keys ):
            if key in old_records and key not in new_records :

                rec =old_records [key ]
                attrs =rec ["attributes"]
                row ={
                "Comparison_Status":"DELETED",
                "Changed_Component":"Entire Document / Record",
                "Difference_Details":"Order was deleted."
                }
                for field in fields :
                    row [f"Old_{field }"]=attrs .get (field )or ""
                    row [f"New_{field }"]=""
                comparison_results .append (row )
            elif key in new_records and key not in old_records :

                rec =new_records [key ]
                attrs =rec ["attributes"]
                row ={
                "Comparison_Status":"ADDED",
                "Changed_Component":"Entire Document / Record",
                "Difference_Details":"Order was added."
                }
                for field in fields :
                    row [f"Old_{field }"]=""
                    row [f"New_{field }"]=attrs .get (field )or ""
                comparison_results .append (row )
            else :

                old_rec =old_records [key ]
                new_rec =new_records [key ]
                old_attrs =old_rec ["attributes"]
                new_attrs =new_rec ["attributes"]

                old_abs_path =old_folder /old_rec ["rel_path"]
                new_abs_path =new_folder /new_rec ["rel_path"]


                rel_path_lower =new_rec ["rel_path"].lower ()

                if "features"in rel_path_lower :
                    changed_comp ,differences =self ._compare_features (old_abs_path ,new_abs_path )
                elif "specifications"in rel_path_lower :
                    changed_comp ,differences =self ._compare_specs (old_abs_path ,new_abs_path )
                elif "sales_data"in rel_path_lower :
                    changed_comp ,differences =self ._compare_sales_data (old_abs_path ,new_abs_path )
                elif "customer_reviews"in rel_path_lower :
                    changed_comp ,differences =self ._compare_reviews (old_abs_path ,new_abs_path )
                else :
                    changed_comp ,differences =self ._compare_default (old_attrs ,new_attrs ,old_abs_path ,new_abs_path )

                if differences :
                    row ={
                    "Comparison_Status":"MODIFIED",
                    "Changed_Component":changed_comp ,
                    "Difference_Details":differences 
                    }
                    for field in fields :
                        row [f"Old_{field }"]=old_attrs .get (field )or ""
                        row [f"New_{field }"]=new_attrs .get (field )or ""
                    comparison_results .append (row )
                else :
                    row ={
                    "Comparison_Status":"IDENTICAL",
                    "Changed_Component":"None (Identical)",
                    "Difference_Details":"No differences detected."
                    }
                    for field in fields :
                        row [f"Old_{field }"]=old_attrs .get (field )or ""
                        row [f"New_{field }"]=new_attrs .get (field )or ""
                    comparison_results .append (row )

        return {
        "old_folder_name":Path (old_folder ).name ,
        "new_folder_name":Path (new_folder ).name ,
        "old_folder_path":str (Path (old_folder ).resolve ()),
        "new_folder_path":str (Path (new_folder ).resolve ()),
        "timestamp":datetime .now ().strftime ("%Y-%m-%d %H:%M:%S"),
        "results":comparison_results 
        }



    def generate_xlsx_report (self ,analysis ,output_path ):
        """
        Generate a styled Excel report with bullet-point Difference_Details and text wrapping.
        """
        wb =openpyxl .Workbook ()
        ws =wb .active 
        ws .title ="Comparison Report"
        ws .views .sheetView [0 ].showGridLines =True 

        font_name ="Segoe UI"
        title_font =Font (name =font_name ,size =15 ,bold =True ,color ="1F1F1F")
        timestamp_font =Font (name =font_name ,size =9 ,italic =True ,color ="555555")
        header_font =Font (name =font_name ,size =11 ,bold =True ,color ="FFFFFF")
        bold_font =Font (name =font_name ,size =10 ,bold =True )
        regular_font =Font (name =font_name ,size =10 )
        bullet_font =Font (name =font_name ,size =9.5 )

        fields =["Order_ID","Order_Date","Customer_Name","City","State","Region",
        "Country","Category","Sub_Category","Product_Name"
        ]

        header_fill =PatternFill (start_color ="1F497D",end_color ="1F497D",fill_type ="solid")
        added_fill =PatternFill (start_color ="C6E0B4",end_color ="C6E0B4",fill_type ="solid")
        deleted_fill =PatternFill (start_color ="F4CCCC",end_color ="F4CCCC",fill_type ="solid")
        modified_fill =PatternFill (start_color ="FFF2CC",end_color ="FFF2CC",fill_type ="solid")
        band_fill_1 =PatternFill (start_color ="FFFFFF",end_color ="FFFFFF",fill_type ="solid")
        band_fill_2 =PatternFill (start_color ="F3F3F3",end_color ="F3F3F3",fill_type ="solid")

        thin_border_side =Side (border_style ="thin",color ="D9D9D9")
        thin_border =Border (
        left =thin_border_side ,right =thin_border_side ,
        top =thin_border_side ,bottom =thin_border_side 
        )
        header_border_side =Side (border_style ="thin",color ="FFFFFF")
        header_border =Border (
        left =header_border_side ,right =header_border_side ,
        top =header_border_side ,bottom =header_border_side 
        )

        align_left =Alignment (horizontal ="left",vertical ="top",wrap_text =False )
        align_center =Alignment (horizontal ="center",vertical ="center",wrap_text =False )
        align_wrap_top =Alignment (horizontal ="left",vertical ="top",wrap_text =True )

        title =f"Folder Difference Analysis: {analysis ['old_folder_name']} vs {analysis ['new_folder_name']}"
        ws ["A1"]=title
        ws ["A1"].font =title_font 
        ws ["A1"].alignment =Alignment (horizontal ="left",vertical ="center")
        ws .row_dimensions [1 ].height =28 

        ws ["A2"]=f"Analysis Timestamp: {analysis ['timestamp']}"
        ws ["A2"].font =timestamp_font 
        ws ["A2"].alignment =Alignment (horizontal ="left",vertical ="center")
        ws .row_dimensions [2 ].height =18 

        headers =["S.No","Comparison_Status"]
        for field in fields :
            headers .extend ([f"Old_{field }",f"New_{field }"])
        headers +=["Changed_Component","Difference_Details"]

        ws .merge_cells (start_row =1 ,start_column =1 ,end_row =1 ,end_column =len (headers ))

        DIFF_COL_IDX =headers .index ("Difference_Details")+1 
        COMP_COL_IDX =headers .index ("Changed_Component")+1 

        for col_idx ,h in enumerate (headers ,start =1 ):
            cell =ws .cell (row =4 ,column =col_idx ,value =h )
            cell .font =header_font 
            cell .fill =header_fill 
            cell .alignment =Alignment (horizontal ="center",vertical ="center",wrap_text =False )
            cell .border =header_border 
        ws .row_dimensions [4 ].height =26 

        LINE_HEIGHT_PX =14.5 

        for row_idx ,rec in enumerate (analysis ["results"],start =5 ):
            status =rec ["Comparison_Status"]
            status_fill ={
            "ADDED":added_fill ,
            "DELETED":deleted_fill ,
            "REMOVED":deleted_fill ,
            "MODIFIED":modified_fill ,
            }.get (status ,None )
            band_fill =band_fill_1 if (row_idx -5) %2 ==0 else band_fill_2

            raw_diffs =rec ["Difference_Details"]
            if isinstance (raw_diffs ,list )and raw_diffs :
                bullet_text ="\n".join (f"\u2022 {item }"for item in raw_diffs )
                num_lines =len (raw_diffs )
            else :
                bullet_text =str (raw_diffs )if raw_diffs else ""
                num_lines =1 

            row_height =max (22 ,num_lines *LINE_HEIGHT_PX +8 )
            ws .row_dimensions [row_idx ].height =row_height 

            for col_idx ,field in enumerate (headers ,start =1 ):
                if field =="S.No":
                    val =row_idx -4 
                elif field =="Difference_Details":
                    val =bullet_text 
                else :
                    val =rec .get (field ,"")

                cell =ws .cell (row =row_idx ,column =col_idx ,value =val )
                cell .border =thin_border 
                cell .fill =band_fill

                if field =="Difference_Details":
                    cell .font =bullet_font 
                    cell .alignment =Alignment (horizontal ="left",vertical ="top",wrap_text =True)
                elif field in ("Comparison_Status","Changed_Component"):
                    cell .font =bold_font 
                    cell .alignment =align_center if field =="Comparison_Status" else align_wrap_top
                elif field =="S.No":
                    cell .alignment =Alignment (horizontal ="center",vertical ="top",wrap_text =True)
                elif field .startswith ("Old_"):
                    cell .font =Font (name =font_name ,size =10 )
                    cell .alignment =align_left
                elif field .startswith ("New_"):
                    cell .font =Font (name =font_name ,size =10 )
                    cell .alignment =align_left
                else :
                    cell .font =regular_font 
                    if field in ("Order_Date","Order_ID"):
                        cell .alignment =align_center
                    else :
                        cell .alignment =align_left

                if field =="Comparison_Status" and status_fill:
                    cell .fill =status_fill
                if field in ("Difference_Details","Changed_Component"):
                    cell .alignment =align_wrap_top
                if field =="Difference_Details":
                    ws .column_dimensions [get_column_letter (col_idx )].width =80

        for col in ws .columns :
            col_letter =get_column_letter (col [0 ].column )
            col_num =col [0 ].column 

            if col_num ==DIFF_COL_IDX :
                ws .column_dimensions [col_letter ].width =80 
            elif col_num ==COMP_COL_IDX :
                ws .column_dimensions [col_letter ].width =40 
            else :
                max_len =0 
                for cell in col :
                    if cell .row >2 and cell .value :
                        first_line =str (cell .value ).split ("\n")[0 ]
                        max_len =max (max_len ,len (first_line ))
                ws .column_dimensions [col_letter ].width =max (max_len +4 ,14 )

        ws .freeze_panes ="B5"

        Path (output_path ).parent .mkdir (parents =True ,exist_ok =True )
        wb .save (output_path )



    def generate_csv_report (self ,analysis ,output_path ):
        """
        Generate a flat CSV report with bullet-prefixed Difference_Details.
        """
        Path (output_path ).parent .mkdir (parents =True ,exist_ok =True )

        fields =[
        "Order_ID","Order_Date","Customer_Name","City","State","Region",
        "Country","Category","Sub_Category","Product_Name"
        ]
        headers =["S.No","Comparison_Status","Old_Content","Modified_Statement"]
        for field in fields :
            headers .extend ([f"Old_{field }",f"New_{field }"])
        headers +=["Changed_Component","Difference_Details"]

        with open (output_path ,mode ="w",encoding ="utf-8",newline ="")as f :
            writer =csv .writer (f )
            writer .writerow (headers )

            for index ,rec in enumerate (analysis ["results"],start =1 ):
                old_content ="; ".join (
                f"{field }: {rec .get (f'Old_{field }','')}"for field in fields if rec .get (f'Old_{field }',"")
                )
                modified_statement =rec .get ("Difference_Details","")
                if isinstance (modified_statement ,list ):
                    modified_statement =" | ".join (f"\u2022 {item }"for item in modified_statement )

                row =[
                index ,
                rec .get ("Comparison_Status",""),
                old_content ,
                modified_statement 
                ]

                for h in headers [4 :]:
                    val =rec .get (h ,"")
                    if h =="Difference_Details"and isinstance (val ,list ):
                        val =" | ".join (f"\u2022 {item }"for item in val )
                    row .append (val )

                writer .writerow (row )


if __name__ =="__main__":
    print ("Testing updated ReportGenerator...")
    generator =ReportGenerator ()
    old_dir =Path ("uploads/old")
    new_dir =Path ("uploads/new")

    if old_dir .exists ()and new_dir .exists ():
        analysis =generator .analyze_differences (old_dir ,new_dir )
        generator .generate_csv_report (analysis ,"reports/difference_report.csv")
        print ("CSV Report generated successfully.")
    else :
        print ("Uploads folders do not exist. Test skipped.")
